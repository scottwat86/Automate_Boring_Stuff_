#! python3

# PP20_text_to_excel.py

# Write a program to read in the contents of several text files (you can make the text files yourself)
# and insert those contents into a spreadsheet, with comma delimited items are inserted on rows
# within one column. 

# By Scott Watson

# Import modules
import os, sys #####
import openpyxl as xl
from openpyxl.utils import get_column_letter

# Command line argument verification
if len(sys.argv) < 2:
    print('Usage: python3 PP20_text_to_excel.py <csv file> ')
    sys.exit(-1)

# Creates list of files to import and checks if they are valid files
import_list = []
not_found_list = []
for i in range(1, len(sys.argv)):
    if not os.path.exists(sys.argv[i]):
        not_found_list.append(sys.argv[i])
        print('\n*********************')
        print(f'File {sys.argv[i]} does not exist')
        print('*********************')
        continue
    import_list.append(sys.argv[i])

# Load valid files into list variable
data_list = []
i = 0
for file in import_list:
    with open(file) as text_file:
        data_list.append('')
        data_list[i] = text_file.read().split(',')
    i += 1

# Creates workbook and selects active sheet
wb = xl.Workbook()
sheet = wb.active

# Cycles through the import_list, inserts the file name in row 1, and each comma delimited file is
# inserted  on a new row
for i in range(len(data_list)):
    sheet[get_column_letter(i + 1) + '1'].value = import_list[i]
    for j in range(len(data_list[i])):
        sheet[get_column_letter(i + 1) + str(j + 2)].value = data_list[i][j]

print('\n*********************')
print(f'Files{import_list} were imported')
print('*********************')
print(f'Files{not_found_list} were not found and NOT imported\n')

print('Saving excel file')
wb.save('text_data.xlsx')

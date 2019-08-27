#! python3

# PP19_Spreadsheet_Cell_Inverter.py

# Write a program to invert the row and column of the cells in the spreadsheet.
# For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
# This should be done for all cells in the spreadsheet.

# By Scott Watson

# Import modules
import sys
import openpyxl as xl
from openpyxl.utils import get_column_letter

# Check if arguments are correct length and if the Excel file exists
if len(sys.argv) < 2:
    print('Usage: python3 PP19_Spreadsheet_Cell_Inverter.py <excel filename>')
    sys.exit(-1)
elif not os.path.exists(sys.argv[1]):
    print(f'File {sys.argv[1]} does not exist.')
    sys.exit(-2)

# Loads the existing input file and creates new output file
input_file = sys.argv[1]
input_wb = xl.load_workbook(input_file)
output_wb = xl.Workbook()
input_sheet = input_wb.active
inverted_sheet = output_wb. active

# Iterates through all the cells with values swapping coloumns with rows and vice versa
for row in range(1, input_sheet.max_row + 1):
    for column in range(1, input_sheet.max_column + 1):
        inverted_sheet [get_column_letter(row) + str(column)].value = \
                                                                input_sheet[get_column_letter(column) + str(row)].value

# Creates a name for the new file based on input file name and then closes files
base_name = input_file.split('.')[0]
output_wb.save('inverted_' + base_name + '.xlsx')
input_wb.close()

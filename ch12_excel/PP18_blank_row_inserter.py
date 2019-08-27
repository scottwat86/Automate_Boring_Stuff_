#! python3

# PP17_blank_row_inserter.py

# Create a program blankRowInserter.py that takes two integers and a filename string as command
# line arguments. Let’s call the first integer N and the second integer M. Starting at row N,
# the program should insert M blank rows into the spreadsheet.

# By Scott Watson

import logging
import openpyxl as xl
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.DEBUG,
                                  format=' %(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)  # Logging disabled

# Initiate variables
if len(sys.argv) == 4:
    row_insertion = int(sys.argv[1])
    number_to_insert = int(sys.argv[2])
    excel_filename = sys.argv[3]
else:
    print('''Command line was invalid. e.g.
            ******************************************

            python3 blankRowInserter.py m n excelFilePath

            ******************************************
            Please try again
            ''')

    sys.exit(-1)

wb = xl.load_workbook(excel_filename)
sheet = wb.active

# Find the extent of the last row and columns
last_row = sheet.max_row
last_column = sheet.max_column

print(f"Shifting rows by {number_to_insert} downward, starting at last row and ending at {row_insertion} row\n")

logging.debug('Looping through cells starting at last row until the insertion row')
for row in range(last_row, row_insertion,-1):
    for column in range(1, last_column+1):
        logging.debug('Shifting cell values')
        sheet[get_column_letter(column) + str(row + number_to_insert)] = \
                                                                            sheet[get_column_letter(column) + str(row)].value

        logging.debug('Checking for boundary conditions at insertion')
        # Starting at the last row and shift everything down by number_to_insert
        if column == last_column and row == row_insertion:
            sheet[get_column_letter(column) + str(row)] = ''
            print(sheet[get_column_letter(column) + str(row)])
            break
        elif row_insertion  <= row <= row_insertion + number_to_insert:
            sheet[get_column_letter(column) + str(row)] = ''

print("Rows have been inserted. Done")

logging.debug('Saving new file')
file = excel_filename.split('.')
wb.save(file[0] + '_2.' + file[1])

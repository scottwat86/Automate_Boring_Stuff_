#               Automate the Boring Stuff with Python 3
#               Chapter 12 – Working with Excel Spreadsheets

# OpenPyXL is a package for reading and writing Excel files. OpenPyXL covers  advanced
# features of Excel such as charts, styles, number formatting and conditional formatting.
# It even includes a tokeniser for parsing Excel formulas and isupports Pandas as well

# Can be quite slow for handling large files -> If you have to write reports with thousands of rows
# and your application is time-sensitive then XlsxWriter or PyExcelerate may be better choices

# More Info for Excel modules
# https://www.pyxll.com/blog/tools-for-working-with-excel-and-python/

'''
I recognize that I will be repeating some import statments and other code but in the interest of
learning I will follow the book literally and retype the code to better remember in the future
'''


import openpyxl, os

# Defines local variable path for Ch12
path = os.environ['python_home'] + '\\Automate_Boring_Stuff_\\ch12_excel'
os.chdir(path)

# Opening workbook examples.xlsx
wb = openpyxl.load_workbook('example.xlsx')
type(wb) # openpyxl.workbook.workbook.Workbook

# GETTOMG SHEETS FROM THE WORKBOOK
import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
# wb.get_sheet_names() DEPRECATED VERSION
wb.sheetnames # ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']

# sheet = wb.get_sheet_by_name('Sheet3') DEPRECATED VERSION
sheet = wb['Sheet3']

sheet # <Worksheet "Sheet3">
type(sheet) #openpyxl.worksheet.worksheet.Worksheet
sheet.title # 'Sheet3'

anotherSheet = wb.active # Sheet1 is the active sheet
anotherSheet #<Worksheet "Sheet1">

# GETTING CELLS FROM THE SHEETS

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1') DEPRECATED VERSION
sheet = wb['Sheet1']

sheet['A1'] #<Cell 'Sheet3'.A1>
sheet['A1'].value #datetime.datetime(2015, 4, 5, 13, 34, 2)
# openpyxl interprets the datetime data instead of a string

c = sheet['B1'] # Initlize variable with openpyxl.cell.cell.Cell object
c.value # 'Apples'

'Row ' + str(c.row) + '. Column ' + c.column_letter + ' is ' + c.value #'Row 1. Column B is Apples'
# cell.column revised to cell.column_letter as column is only the int index not the letter

'Cell ' + c.coordinate + ' is ' + c.value # 'Cell B1 is Apples'
sheet['C1'].value # 73

# First row / column starts at 1 NOT 0
sheet.cell(row=1, column=2) #<Cell 'Sheet1'.B1>
sheet.cell(row=1, column=2).value # 'Apples'

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
'''
1 Apples
3 Pears
5 Apples
7 Strawberries
'''

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1')
sheet = wb['Sheet1']
sheet.max_row # 7
sheet.max_column # 3 -> C

# CONVERTING BETWEEN COLUMN LETTERS AND NUMBERS
import openpyxl

# This code is out of date  openpyxl.cell became -> openpyxl.util
from openpyxl.utils import get_column_letter, column_index_from_string

# now its just get_column
get_column_letter(1) # 'A'
get_column_letter(2) # 'B'
get_column_letter(27) # 'AA'
get_column_letter(900) # 'AHP'

wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1') DEPRECATED VERSION
sheet = wb['Sheet1']
get_column_letter(sheet.max_column) # 'C' is the largest column with a value in it

column_index_from_string('A') # 1
column_index_from_string('AA') # 27

# GETTING ROWS AND COLUMNS FROM THE SHEETS
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1') DEPRECATED VERSION
sheet = wb['Sheet1']
tuple(sheet['A1' : 'C3'])
'''
((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>),
 (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
 (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
'''

for rowOfCellObjects in sheet['A1' : 'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

'''
A1 2015-04-05 13:34:02
B1 Apples
C1 73
--- END OF ROW ---
A2 2015-04-05 03:41:23
B2 Cherries
C2 85
--- END OF ROW ---
A3 2015-04-06 12:46:51
B3 Pears
C3 14
--- END OF ROW ---

'''
# sheet['A1' : 'C3']   -> provies a geneterator object of the cells included in the range

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')

# Revise sheet.columns[1] -> sheet['B'] now
sheet['B']
'''
(<Cell 'Sheet1'.B1>,
 <Cell 'Sheet1'.B2>,
 <Cell 'Sheet1'.B3>,
 <Cell 'Sheet1'.B4>,
 <Cell 'Sheet1'.B5>,
 <Cell 'Sheet1'.B6>,
 <Cell 'Sheet1'.B7>)
'''

for cellObj in sheet['B']:
    print(cellObj.value)
'''
Apples
Cherries
Pears
Oranges
Apples
Bananas
Strawberries
'''

# WORKBOOKS, SHEETS, CELLS
import openpyxl # Import module
wb = openpyxl.load_workbook('example.xlsx') # Get Workbook object
active_sheet = wb.active # Read Active Sheet Variable
active_sheet.title # 'Sheet1'
# sheet = wb.get_sheet_by_name('Sheet1') DEPRECATED VERSION
sheet = wb['Sheet1'] # Retreive Sheet object

sheet.cell(row=1, column=2) # <Cell 'Sheet1'.B1>  cell()
sheet['B1'] # <Cell 'Sheet1'.B1>  indexing

#############################
#Project: Reading Data from a Spreadsheet
# See  readCensusExcel.py.
#############################

# WRITING EXCEL DOCUMENTS

import openpyxl
wb = openpyxl.Workbook()
# wb.get_sheet_names()  DEPRECATED VERSION
wb.sheetnames
sheet = wb.active # Normally active sheet is index=0
sheet.title # 'Sheet'
sheet.title = 'Spam Bacon Eggs Sheet'
wb.sheetnames

import openpyxl
wb = openpyxl.load_workbook('example.xlsx') # Opens the file
sheet = wb.active # openpyxl.worksheet.worksheet.Worksheet
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')

# CREATING AND REMOVING SHEETS
import openpyxl
wb = openpyxl.Workbook()
# wb.get_sheet_names()    DEPRECATED VERSION
wb.sheetnames # ['Sheet']
wb.create_sheet() # <Worksheet "Sheet1">
wb.sheetnames
wb.create_sheet(index=0, title='First Sheet') # <Worksheet "First Sheet">
wb.sheetnames
wb.create_sheet(index=2, title='Middle Sheet')
wb.sheetnames  # ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
wb.remove(wb['Middle Sheet']) # THIS IS THE NEW WAY TO DO IT
# wb.remove(wb.get_sheet_by_name('Middle Sheet'))      DEPRECATED VERSION
wb.remove(wb['Sheet1'])
# wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))     DEPRECATED VERSION
wb.sheetnames  #['First Sheet', 'Sheet']

# WRITING VALUES TO CELLS
import openpyxl
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello world!'
sheet['A1'].value # 'Hello world!'

#############################
# Project: Updating a Spreadsheet
# See  updateProduce.py
#############################


# SETTING THE FONT STYLE OF CELLS

import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # openpyxl.styles.fonts.Font object
sheet['A1'].font = italic24Font # assigning the font object ot the cell font attribute
sheet['A1'] = 'Hello world!'
wb.save('styled.xlsx')

# FONT OBJECTS
'''
KEYWORD ARG                 DATA TYPE             DECRIPTION
name                                       String                        font name, eg 'Calibri' or 'Times New Roman'
size                                          Integer                     point size
bold                                         Boolean                   True for bold font
italic                                        Boolean                   True for italic font
'''

import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']

fontObj = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj
sheet['A1'] = 'Bold Time New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

wb.save('styles.xlsx')

# FORMULAS
# Formulas, which begin with an equal sign, can configure cells to contain values calculated from other cells.

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')

# SETTING ROW HEIGHT AND COLUMN WIDTH
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

# MERGING AND UNMERGING CELLS
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

# UNMERGE
import openpyxl
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')

# FREEZE PANES
# In OpenPyXL, each Worksheet object has a freeze_panes attribute that can be set to a
# Cell object or a string of a cell’s coordinates

'''
FREEZE_PANES SETTING                                    ROW / COLUMNS FROZEN

sheet.freeze_panes = 'A2'                                          Row 1
sheet.freeze_panes = 'B1'                                          Column A
sheet.freeze_panes = 'C1'                                          Columns A and B
sheet.freeze_panes = 'C2'                                          Row 1 and columns A and B
sheet.freeze_panes = 'A1' or                                      No frozen panes
                sheet.freeze_panes = None
'''

import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')

# CHARTS
# OpenPyXL supports creating bar, line, scatter, and pie charts using the data in a sheet’s cells.

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, 11): # Create some data in the column A
    sheet['A'+str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1,  max_col=1, max_row=10)
# min_col / max_col / min_row / max_row defines limits of data to be graphed under First Series
# Note FIRST ROW = 1    NOT 0
'''
Reference(sheet, min_col=1, min_row=1,  max_col=1, max_row=10)  contains 2 pairs of tuples

(row,col),   (row,col)
(1, 1),          (10, 1)             ------> A1:A10
(3, 2),          (6, 4)               ------> B3:C6
(5, 3),          (5, 3)               ------> C5:C5
'''
seriesObj = openpyxl.chart.Series(refObj, title='First series')
# Creates a Series object that selects the cells defined in the Ref object

chartObj = openpyxl.chart.BarChart()  # Creates a Chart object of type BarChart
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5') # Inserts the Chart Object at cell C5

wb.save('sampleChart.xlsx')

#Practice Questions
#   1)What does the openpyxl.load_workbook() function return?
openpyxl.load_workbook('file.xlsx') # returns a saved excel file

#   2)What does the get_sheet_names() workbook method return?
#   wb.get_sheet_name() is deprecated;
wb.sheetnames # is the prefered way and it returns sheet names in  the workbook

#   3) How would you retrieve the Worksheet object for a sheet named 'Sheet1'?
#    wb.get_sheet_by_name('Sheet1') is deprecated and
wb['Sheet1'] # prefered way

#   4) How would you retrieve the Worksheet object for the workbook’s active sheet?
wb.active

#   5) How would you retrieve the value in the cell C5?
sheet['C5'].value # prefered

#   6) How would you set the value in the cell C5 to "Hello"?
sheet['C5'] = 'Hello'

#   7) How would you retrieve the cell’s row and column as integers?
sheet.row
sheet.column

#   8) What do the max_column and max_row sheet methods return, and what is the data type of these return values?
sheet.max_row
sheet.max_column
# max_row/column returns the integer index of the  maximum extent of data in the sheet

#   9) If you needed to get the integer index for column 'M', what function would you need to call?
from openpyxl.utils import column_index_from_string
column_index_from_string('M')
# OR
openpyxl.cell.column_index_from_string('M')

#   10) If you needed to get the string name for column 14, what function would you need to call?
openpyxl.cell.get_column_letter(14)

#   11) How can you retrieve a tuple of all the Cell objects from A1 to F1?
sheet['A1' : 'F1']

#   12) How would you save the workbook to the filename example.xlsx?
wb.save('example.xlsx')

#   13) How do you set a formula in a cell?
sheet['A1'] = '=SUM(A1:A2)'

#   14) How would you set the height of row 5 to 100?
sheet.row_dimensions[5].height = 100

#   15) How would you hide column C?
sheet.column_dimensions['C'].hidden = True

#   16) Name a few features that OpenPyXL 2.3.3 does not load from a spreadsheet file.
# Doesn't load charts, print titles, images, freeze panes

#   17) What is a freeze pane?
# Freeze pane allows you to always show a row or column; its good for labels

#   18) What five functions and methods do you have to call to create a bar chart?
openpyxl.charts.Reference()
openpyxl.charts.Series()
openpyxl.charts.BarChart()
chartObj.append(seriesObj)
add_chart()

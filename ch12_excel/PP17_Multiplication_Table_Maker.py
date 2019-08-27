#! python3

# PP17_Multiplication_Table_Maker.py
# Create a program multiplicationTable.py that takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet.

# By Scott Watson

import logging, sys
import openpyxl as xl
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.DEBUG,
                                  format=' %(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)  # Logging disabled

# Initiate variables
if len(sys.argv) == 4:
    row_insertion = int(sys.argv[1])
    number_to_insert = int(sys.argv[2])
    excel_filename = sys.argv[3]
else:
    print('''Program takes row number to insert new row, number of inserted rows, and output excel
    filename as input. The arguments provided either too many or too few to complete the progran.''')
    print('Please try again')
    quit()

print("Filling new Excel Workbook with multiplication table")

wb = xl.Workbook()
sheet = wb.active

row_n = {}
col_n = {}
ref_col = 'A'
ref_row = '2'

sheet.merge_cells('A1:' + get_column_letter(n+2) + '1' )
title_cell = sheet['A1']
title_cell.value = f'Multiplication Table {n} x {n}'
title_cell.font = xl.styles.Font(size=20)
# title_cell.alignment = xl.styles.Alignment(horizontal='center')  # These aren't working
# title_cell.alignment = xl.styles.Alignment(vertical='center')
# title_cell.font = xl.styles.Font(bold=True)

for i in range(2, n+2):
    """ Populates the worksheet with a N length reference column and row skipping first cell A1"""
    ref_col_cell = ref_col  + str(i) # Ref Column is static, only rows change
    ref_row_cell = get_column_letter(i) + ref_row # Ref Rowsis static, only Columns change

    # range begins at 2 therefore -1 to begin count at 1
    sheet[ref_col_cell] = i - 1
    sheet[ref_row_cell] = i - 1

    print('.', end='')# Processing .
    logging.debug('Ref Col Cell ={}\nRef Row Cell ={}'.format(ref_col_cell, ref_row_cell))

offset = 3
for column in range(2, n+2):
    """
    Cycles through the columns and rows and inserting a formula to calculate the multiplication table

    Example of sequence:
    active_cell = Ref Row * Ref Col
    B2.....Bn+2      B=2      B2 = A2 * B1       B3 = A3 * B1 ...
    C2....Bn+2       C=3      C2 = A2 * C1 ...
    .
    Z2....Zn+2        Z=n      Z2 = A2 * Z1 ...
    """
    print('.', end='') # Processing .

    current_column = get_column_letter(column)
    current_ref_col = current_column + '1' # Reference column of the multiplication table

    for row in range(offset, n + 2):
        # current_row = row
        active_cell = current_column + str(row)
        current_ref_row = current_column + ref_row
        current_ref_col = ref_col + str(row)

        sheet[active_cell] = "=" + current_ref_row + "*" + current_ref_col
        logging.debug(f'active_cell ={active_cell}\ncurrent_ref_row = {current_ref_row}\n'  +
                                                                                                f'current_ref_col = {current_ref_col}')

print("\nComplete\nGoodbye")
wb.save('mulitplaction_table.xlsx')
# logging.debug('Value of column {}'.format(col_letter + str(row)))

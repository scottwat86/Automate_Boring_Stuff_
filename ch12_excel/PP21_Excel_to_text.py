#! python3

# PP21_Excel_to_text.py

# Write a program that performs the tasks of the previous program in reverse order:
# The program should open a spreadsheet and write the cells of column A into one text file, the
# cells of column B into another text file, and so on.

# By Scott Watson

# Import modules
import sys
import openpyxl as xl
from openpyxl.utils import get_column_letter

# Command line argument verification
if len(sys.argv) < 2:
    print('Usage: python3 PP21_Excel_to_text.py <excel_1> .. .<excel_n>')
    sys.exit(-1)

print(sys.argv[1])
wb = xl.load_workbook(sys.argv[1])
sheet = wb.active

for column in range(1, sheet.max_column + 1):
    file_name = 'Column_' + get_column_letter(column) + '_.txt'

    with open(file_name, 'a') as file:
        for row in range(1, sheet.max_row + 1):
            cell = sheet[get_column_letter(column)+str(row)].value
            if cell == None:
                continue
            file.write(str(cell) + ',')
